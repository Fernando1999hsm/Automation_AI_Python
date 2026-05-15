import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Casos de Prueba - Saucedemo"

headers = ["ID", "Tipo", "Descripcion", "Pasos", "Datos para probar", "Resultado esperado", "Prioridad"]

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
wrap_align = Alignment(vertical="top", wrap_text=True)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

test_cases = [
    # ==================== LOGIN - POSITIVOS ====================
    [
        "TC-LOG-001",
        "Positivo",
        "Login exitoso con usuario standard_user",
        "1. Navegar a https://www.saucedemo.com/\n2. En el campo 'Username' ingresar el usuario\n3. En el campo 'Password' ingresar la contrasena\n4. Hacer clic en el boton 'Login'\n5. Verificar redireccion a /inventory.html",
        "Usuario: standard_user\nPassword: secret_sauce",
        "Redirecciona a /inventory.html. Se muestra el catalogo de productos con 6 items. El titulo de la pagina es 'Products'. El boton de login ya no esta visible.",
        "Alta"
    ],
    [
        "TC-LOG-002",
        "Positivo",
        "Login exitoso con usuario problem_user y verificacion de imagenes",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'problem_user' en Username\n3. Ingresar 'secret_sauce' en Password\n4. Clic en Login\n5. Verificar que las imagenes de los productos se cargan (aunque sean identicas todas ellas)",
        "Usuario: problem_user\nPassword: secret_sauce",
        "Login exitoso. Las imagenes de productos se cargan (pueden ser la misma imagen repetida). El resto de la funcionalidad (add to cart, checkout) opera correctamente.",
        "Alta"
    ],
    [
        "TC-LOG-003",
        "Positivo",
        "Login exitoso con usuario performance_glitch_user y medicion de tiempo",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'performance_glitch_user' en Username\n3. Ingresar 'secret_sauce' en Password\n4. Clic en Login\n5. Medir el tiempo desde clic hasta que /inventory.html carga completamente",
        "Usuario: performance_glitch_user\nPassword: secret_sauce",
        "Login exitoso. La pagina de inventario carga completa aunque con un retraso notable (2-5 segundos adicionales comparado con standard_user). Todos los elementos son interactivos despues de la carga.",
        "Media"
    ],
    [
        "TC-LOG-004",
        "Positivo",
        "Login exitoso con usuario error_user y validacion de errores en formularios",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'error_user' en Username\n3. Ingresar 'secret_sauce' en Password\n4. Clic en Login\n5. Ir al carrito y proceder al checkout\n6. Intentar completar checkout con datos validos",
        "Usuario: error_user\nPassword: secret_sauce\nCheckout: Nombre, Apellido, CP validos",
        "Login exitoso. Durante el checkout pueden aparecer errores especificos en campos del formulario (el usuario error_user introduce errores intencionales en campos). Verificar que los mensajes de error son coherentes.",
        "Media"
    ],
    [
        "TC-LOG-005",
        "Positivo",
        "Login exitoso con usuario visual_user y verificacion de layout",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'visual_user' en Username\n3. Ingresar 'secret_sauce' en Password\n4. Clic en Login\n5. Verificar que todos los elementos UI se renderizan correctamente (botones, imagenes, textos)",
        "Usuario: visual_user\nPassword: secret_sauce",
        "Login exitoso. Los elementos visuales pueden tener diferencias de posicionamiento/tamano comparado con standard_user. Verificar que todos los elementos son funcionales aunque su apariencia sea distinta.",
        "Media"
    ],

    # ==================== LOGIN - NEGATIVOS ====================
    [
        "TC-LOG-006",
        "Negativo",
        "Login con credenciales invalidas",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'usuario_invalido' en Username\n3. Ingresar 'password_invalido' en Password\n4. Clic en Login\n5. Verificar mensaje de error",
        "Usuario: usuario_invalido\nPassword: password_invalido",
        "Aparece mensaje de error: 'Username and password do not match any user in this service'. NO redirecciona a /inventory.html. El campo Username mantiene el valor ingresado. El campo Password se limpia.",
        "Alta"
    ],
    [
        "TC-LOG-007",
        "Negativo",
        "Login con usuario valido y password incorrecto",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'standard_user' en Username\n3. Ingresar 'wrong_password' en Password\n4. Clic en Login\n5. Verificar mensaje de error",
        "Usuario: standard_user\nPassword: wrong_password",
        "Aparece mensaje de error: 'Username and password do not match any user in this service'. Permanece en la pagina de login. El icono de error (X roja) aparece en los campos.",
        "Alta"
    ],
    [
        "TC-LOG-008",
        "Negativo",
        "Login con campos vacios",
        "1. Navegar a https://www.saucedemo.com/\n2. Dejar el campo Username vacio\n3. Dejar el campo Password vacio\n4. Clic en Login\n5. Verificar mensaje de error",
        "Usuario: (vacio)\nPassword: (vacio)",
        "Aparece mensaje de error: 'Username is required'. Permanece en pagina de login. No se realiza ninguna redireccion.",
        "Alta"
    ],
    [
        "TC-LOG-009",
        "Negativo",
        "Login con solo Username lleno y Password vacio",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'standard_user' en Username\n3. Dejar Password vacio\n4. Clic en Login\n5. Verificar mensaje de error",
        "Usuario: standard_user\nPassword: (vacio)",
        "Aparece mensaje de error: 'Password is required'. Permanece en pagina de login.",
        "Alta"
    ],
    [
        "TC-LOG-010",
        "Negativo",
        "Login con usuario locked_out_user",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar 'locked_out_user' en Username\n3. Ingresar 'secret_sauce' en Password\n4. Clic en Login\n5. Verificar mensaje de error especifico",
        "Usuario: locked_out_user\nPassword: secret_sauce",
        "Aparece mensaje de error: 'Sorry, this user has been locked out.' Permanece en la pagina de login. NO hay acceso al inventario.",
        "Alta"
    ],
    [
        "TC-LOG-011",
        "Negativo",
        "Login con SQL Injection en Username",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar \"' OR 1=1 --\" en Username\n3. Ingresar 'secret_sauce' en Password\n4. Clic en Login\n5. Verificar que NO se permite el acceso",
        "Usuario: ' OR 1=1 --\nPassword: secret_sauce",
        "NO debe permitir el acceso. Debe mostrar mensaje de error de credenciales invalidas. El sistema no debe ser vulnerable a inyeccion SQL.",
        "Alta"
    ],
    [
        "TC-LOG-012",
        "Negativo",
        "Login con XSS en campos (Cross-Site Scripting)",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar '<script>alert(1)</script>' en Username\n3. Ingresar cualquier password\n4. Clic en Login\n5. Verificar que NO se ejecuta el script (sin alert)",
        "Usuario: <script>alert(1)</script>\nPassword: x",
        "El script no debe ejecutarse. Debe mostrar mensaje de error normal o sanitizar la entrada. No debe aparecer ninguna alerta del navegador.",
        "Media"
    ],
    [
        "TC-LOG-013",
        "Negativo",
        "Login con caracteres especiales y largos en campos",
        "1. Navegar a https://www.saucedemo.com/\n2. Ingresar cadena de 500 caracteres en Username\n3. Ingresar password normal\n4. Clic en Login\n5. Verificar comportamiento del sistema",
        "Usuario: 'A'*500\nPassword: secret_sauce",
        "El sistema debe manejar la entrada larga sin romperse (crash). Debe mostrar mensaje de error apropiado o truncar la entrada. No debe haber errores 500 ni pantallas en blanco.",
        "Media"
    ],

    # ==================== INVENTARIO / PRODUCTOS ====================
    [
        "TC-INV-001",
        "Positivo",
        "Visualizar catalogo de productos correctamente",
        "1. Login como standard_user\n2. Verificar que se muestran exactamente 6 productos\n3. Verificar que cada producto tiene: imagen, titulo, descripcion, precio y boton Add to cart",
        "Usuario: standard_user\nPassword: secret_sauce",
        "Se muestran 6 productos. Cada producto tiene: imagen visible, nombre enlace, descripcion textual, precio en formato '$XX.XX' y boton 'Add to cart' rojo.",
        "Alta"
    ],
    [
        "TC-INV-002",
        "Positivo",
        "Ordenar productos por nombre A-Z (default)",
        "1. Login como standard_user\n2. Verificar que el selector de orden tiene 'Name (A to Z)' seleccionado\n3. Verificar que los productos aparecen en orden alfabetico A-Z",
        "Usuario: standard_user\nOrden: Name (A to Z)",
        "Los productos se muestran ordenados alfabeticamente de la A a la Z. El primer producto es 'Sauce Labs Backpack'. El ultimo es 'Test.allTheThings() T-Shirt (Red)'.",
        "Alta"
    ],
    [
        "TC-INV-003",
        "Positivo",
        "Ordenar productos por nombre Z-A",
        "1. Login como standard_user\n2. Seleccionar 'Name (Z to A)' en el filtro de orden\n3. Verificar orden descendente por nombre",
        "Usuario: standard_user\nOrden: Name (Z to A)",
        "Los productos aparecen en orden alfabetico inverso (Z-A). El primer producto es 'Test.allTheThings() T-Shirt (Red)'. El ultimo es 'Sauce Labs Backpack'.",
        "Alta"
    ],
    [
        "TC-INV-004",
        "Positivo",
        "Ordenar productos por precio Low to High",
        "1. Login como standard_user\n2. Seleccionar 'Price (low to high)' en el filtro\n3. Verificar orden ascendente por precio",
        "Usuario: standard_user\nOrden: Price (low to high)",
        "Productos ordenados por precio ascendente. El primer producto debe ser el de menor precio ($7.99: Sauce Labs Onesie). El ultimo el de mayor precio ($49.99: Sauce Labs Fleece Jacket).",
        "Alta"
    ],
    [
        "TC-INV-005",
        "Positivo",
        "Ordenar productos por precio High to Low",
        "1. Login como standard_user\n2. Seleccionar 'Price (high to low)' en el filtro\n3. Verificar orden descendente por precio",
        "Usuario: standard_user\nOrden: Price (high to low)",
        "Productos ordenados por precio descendente. El primer producto debe ser el de mayor precio ($49.99: Sauce Labs Fleece Jacket). El ultimo el de menor precio ($7.99: Sauce Labs Onesie).",
        "Alta"
    ],
    [
        "TC-INV-006",
        "Positivo",
        "Agregar producto individual al carrito desde inventario",
        "1. Login como standard_user\n2. Hacer clic en 'Add to cart' del primer producto (Sauce Labs Backpack)\n3. Verificar que el boton cambia a 'Remove'\n4. Verificar que el icono del carrito muestra '1'\n5. Verificar que el producto aparece en /cart.html",
        "Usuario: standard_user\nProducto: Sauce Labs Backpack",
        "El boton del producto cambia a 'Remove' (color rojo oscuro). El icono del carrito superior derecho muestra el badge '1'. Al abrir el carrito, el producto aparece listado con su precio.",
        "Alta"
    ],
    [
        "TC-INV-007",
        "Positivo",
        "Agregar multiples productos al carrito",
        "1. Login como standard_user\n2. Agregar 3 productos al carrito (Backpack, Bike Light, T-Shirt)\n3. Verificar que el badge del carrito muestra '3'\n4. Abrir carrito y verificar los 3 items",
        "Usuario: standard_user\nProductos: Sauce Labs Backpack, Sauce Labs Bike Light, Sauce Labs Bolt T-Shirt",
        "Badge del carrito muestra '3'. La pagina del carrito lista exactamente 3 productos con sus nombres, precios, cantidades y botones 'Remove'.",
        "Alta"
    ],
    [
        "TC-INV-008",
        "Positivo",
        "Remover producto del carrito desde inventario",
        "1. Login como standard_user\n2. Agregar Sauce Labs Backpack al carrito\n3. Hacer clic en 'Remove'\n4. Verificar que el badge del carrito desaparece\n5. Verificar que el boton vuelve a 'Add to cart'",
        "Usuario: standard_user\nProducto: Sauce Labs Backpack",
        "Badge del carrito desaparece (o muestra 0). El boton del producto vuelve a 'Add to cart'. El producto ya no aparece en /cart.html.",
        "Alta"
    ],
    [
        "TC-INV-009",
        "Positivo",
        "Navegar al detalle de producto haciendo clic en el nombre",
        "1. Login como standard_user\n2. Hacer clic en el nombre 'Sauce Labs Backpack'\n3. Verificar pagina de detalle del producto",
        "Usuario: standard_user\nProducto: Sauce Labs Backpack",
        "Redirecciona a /inventory-item.html?id=4. Muestra: imagen grande, nombre, descripcion detallada, precio y boton 'Add to cart'. Hay boton 'Back to products' para regresar.",
        "Alta"
    ],
    [
        "TC-INV-010",
        "Positivo",
        "Agregar producto al carrito desde pagina de detalle",
        "1. Login como standard_user\n2. Clic en nombre del producto para ir al detalle\n3. Clic en 'Add to cart' en la pagina de detalle\n4. Verificar que badge aparece\n5. Clic en 'Back to products' y verificar que el boton en inventario tambien cambio",
        "Usuario: standard_user\nProducto: Sauce Labs Backpack",
        "Badge del carrito aparece. Al volver al inventario, el boton de ese producto muestra 'Remove'. El estado del carrito es consistente entre paginas.",
        "Alta"
    ],
    [
        "TC-INV-011",
        "Negativo",
        "Verificar comportamiento con usuario problem_user en inventario",
        "1. Login como problem_user\n2. Observar las imagenes de los productos en el inventario\n3. Agregar un producto al carrito\n4. Removerlo\n5. Verificar si hay comportamientos erraticos",
        "Usuario: problem_user\nPassword: secret_sauce",
        "Las imagenes de los productos pueden ser todas identicas (problema conocido). Las acciones de add/remove y ordenamiento pueden fallar o comportarse erraticamente. Documentar cualquier desviacion del comportamiento esperado de standard_user.",
        "Media"
    ],

    # ==================== CARRITO ====================
    [
        "TC-CART-001",
        "Positivo",
        "Visualizar carrito vacio",
        "1. Login como standard_user\n2. Hacer clic en el icono del carrito (sin agregar productos)\n3. Verificar pagina del carrito",
        "Usuario: standard_user\nSin productos en carrito",
        "Pagina del carrito muestra: titulo 'Your Cart', boton 'Continue Shopping', boton 'Checkout'. Lista de items vacia. No hay productos mostrados.",
        "Alta"
    ],
    [
        "TC-CART-002",
        "Positivo",
        "Remover producto desde la pagina del carrito",
        "1. Login como standard_user\n2. Agregar 2 productos al carrito\n3. Navegar a /cart.html\n4. Hacer clic en 'Remove' de uno de los productos\n5. Verificar que solo ese producto se elimina",
        "Usuario: standard_user\nProductos: Backpack y Bike Light\nRemover: Backpack",
        "El producto removido desaparece de la lista del carrito. El badge del carrito se actualiza a '1'. El otro producto permanece en la lista. El precio total se actualiza (si hay campo total).",
        "Alta"
    ],
    [
        "TC-CART-003",
        "Positivo",
        "Continuar comprando desde el carrito",
        "1. Login como standard_user\n2. Agregar producto al carrito\n3. Ir al carrito\n4. Clic en 'Continue Shopping'\n5. Verificar que se vuelve al inventario",
        "Usuario: standard_user",
        "Redirecciona a /inventory.html. El badge del carrito mantiene los items previamente agregados. Se puede seguir agregando productos.",
        "Alta"
    ],

    # ==================== CHECKOUT ====================
    [
        "TC-CHK-001",
        "Positivo",
        "Checkout exitoso con datos validos",
        "1. Login como standard_user\n2. Agregar Sauce Labs Backpack al carrito\n3. Ir al carrito\n4. Clic en 'Checkout'\n5. Ingresar nombre, apellido y codigo postal\n6. Clic en 'Continue'\n7. Verificar pagina de overview\n8. Clic en 'Finish'\n9. Verificar pagina de confirmacion",
        "Usuario: standard_user\nNombre: Juan\nApellido: Perez\nCP: 12345",
        "Paso 5: Formulario se llena correctamente. Paso 6: Redirecciona a /checkout-step-two.html con resumen del pedido (item, descripcion, precio, total con tax). Paso 8: Redirecciona a /checkout-complete.html con mensaje 'Thank you for your order!' y checkmark. Paso 9: Aparece boton 'Back Home'.",
        "Alta"
    ],
    [
        "TC-CHK-002",
        "Positivo",
        "Checkout con multiple productos y verificar calculo de total",
        "1. Login como standard_user\n2. Agregar 3 productos de diferentes precios\n3. Ir al carrito\n4. Checkout con datos validos\n5. En overview verificar: subtotal = suma de precios, tax = 8% del subtotal, total = subtotal + tax",
        "Usuario: standard_user\nProductos: Backpack ($29.99), Bike Light ($9.99), Onesie ($7.99)\nSubtotal esperado: $47.97\nTax: $3.84\nTotal: $51.81",
        "La pagina de overview muestra: 'Item total: $47.97', 'Tax: $3.84', 'Total: $51.81'. Los calculos deben ser matematicamente correctos (tax = subtotal * 0.08 redondeado a 2 decimales).",
        "Alta"
    ],
    [
        "TC-CHK-003",
        "Positivo",
        "Checkout completo y regresar al inicio",
        "1. Login como standard_user\n2. Agregar producto, hacer checkout completo\n3. En pagina de confirmacion, clic en 'Back Home'\n4. Verificar que el carrito esta vacio y se vuelve a inventory",
        "Usuario: standard_user",
        "Redirecciona a /inventory.html. El badge del carrito NO aparece (vacio). Es como si se hubiera iniciado una nueva sesion de compra. Los botones de todos los productos muestran 'Add to cart'.",
        "Alta"
    ],
    [
        "TC-CHK-004",
        "Negativo",
        "Checkout con campos obligatorios vacios",
        "1. Login como standard_user\n2. Agregar producto y proceder a checkout\n3. Dejar todos los campos vacios\n4. Clic en 'Continue'\n5. Verificar mensaje de error",
        "Usuario: standard_user\nNombre: (vacio)\nApellido: (vacio)\nCP: (vacio)",
        "Aparece mensaje de error: 'First Name is required'. Permanece en /checkout-step-one.html. NO se avanza a la pagina de overview.",
        "Alta"
    ],
    [
        "TC-CHK-005",
        "Negativo",
        "Checkout con solo First Name y demas vacios",
        "1. Login como standard_user\n2. Agregar producto, ir a checkout\n3. Ingresar solo First Name\n4. Dejar Last Name y Postal Code vacios\n5. Clic en Continue y verificar error",
        "Usuario: standard_user\nNombre: Juan\nApellido: (vacio)\nCP: (vacio)",
        "Aparece mensaje de error: 'Last Name is required'. No se avanza.",
        "Alta"
    ],
    [
        "TC-CHK-006",
        "Negativo",
        "Checkout con First Name y Last Name pero sin Postal Code",
        "1. Login como standard_user\n2. Agregar producto, ir a checkout\n3. Ingresar First Name y Last Name\n4. Dejar Postal Code vacio\n5. Clic en Continue y verificar error",
        "Usuario: standard_user\nNombre: Juan\nApellido: Perez\nCP: (vacio)",
        "Aparece mensaje de error: 'Postal Code is required'. No se avanza.",
        "Alta"
    ],
    [
        "TC-CHK-007",
        "Negativo",
        "Checkout cancelado desde el formulario",
        "1. Login como standard_user\n2. Agregar producto, ir a checkout\n3. Llenar datos\n4. Clic en 'Cancel'\n5. Verificar redireccion",
        "Usuario: standard_user\nDatos: cualquier valor valido",
        "Redirecciona al carrito (/cart.html). Los items en el carrito se mantienen. NO se pierde la seleccion de productos.",
        "Media"
    ],
    [
        "TC-CHK-008",
        "Negativo",
        "Checkout cancelado desde la pagina de overview",
        "1. Login como standard_user\n2. Agregar producto, checkout completo hasta overview\n3. En overview, clic en 'Cancel'\n4. Verificar redireccion",
        "Usuario: standard_user",
        "Redirecciona a /inventory.html. Los items NO se pierden pero se vuelve al inventario (comportamiento: los items quedan en el carrito, no se limpia).",
        "Media"
    ],
    [
        "TC-CHK-009",
        "Negativo",
        "Checkout sin productos en el carrito",
        "1. Login como standard_user\n2. Ir directamente a /cart.html (sin agregar productos)\n3. Verificar si el boton Checkout esta habilitado\n4. Si esta habilitado, hacer clic y verificar comportamiento",
        "Usuario: standard_user\nSin productos en carrito",
        "El boton 'Checkout' debe estar habilitado. Si se hace clic, debe permitir llenar datos y finalizar la compra (sin productos). Alternativamente, si la app bloquea esto, debe mostrar un mensaje indicando que no hay productos.",
        "Media"
    ],

    # ==================== BURGER MENU / NAVEGACION ====================
    [
        "TC-MENU-001",
        "Positivo",
        "Abrir y cerrar burger menu",
        "1. Login como standard_user\n2. Clic en las 3 lineas (burger menu) esquina superior izquierda\n3. Verificar que el menu se despliega\n4. Hacer clic en la X para cerrar\n5. Verificar que el menu se cierra",
        "Usuario: standard_user",
        "Menu se despliega desde la izquierda con opciones: 'All Items', 'About', 'Logout', 'Reset App State'. Al hacer clic en X, el menu se desliza hacia la izquierda y desaparece.",
        "Media"
    ],
    [
        "TC-MENU-002",
        "Positivo",
        "Menu - All Items desde pagina de detalle",
        "1. Login como standard_user\n2. Ir al detalle de un producto\n3. Abrir burger menu\n4. Clic en 'All Items'\n5. Verificar redireccion",
        "Usuario: standard_user\nRuta actual: /inventory-item.html?id=4",
        "Redirecciona a /inventory.html. Muestra todos los productos. No hay errores de navegacion.",
        "Media"
    ],
    [
        "TC-MENU-003",
        "Positivo",
        "Menu - About redirecciona a Sauce Labs",
        "1. Login como standard_user\n2. Abrir burger menu\n3. Clic en 'About'\n4. Verificar redireccion externa",
        "Usuario: standard_user",
        "Redirecciona a https://saucelabs.com/ (sitio externo). La pagina de Sauce Labs se abre en la misma pestana.",
        "Media"
    ],
    [
        "TC-MENU-004",
        "Positivo",
        "Menu - Reset App State",
        "1. Login como standard_user\n2. Agregar 2 productos al carrito\n3. Abrir burger menu\n4. Clic en 'Reset App State'\n5. Verificar que el carrito se vacia y botones vuelven a 'Add to cart'",
        "Usuario: standard_user\nProductos en carrito: 2",
        "El badge del carrito desaparece. Todos los botones de productos vuelven a 'Add to cart'. El estado de la aplicacion se reinicia sin recargar la pagina.",
        "Alta"
    ],
    [
        "TC-MENU-005",
        "Positivo",
        "Menu - Logout",
        "1. Login como standard_user\n2. Abrir burger menu\n3. Clic en 'Logout'\n4. Verificar redireccion a pagina de login",
        "Usuario: standard_user",
        "Redirecciona a https://www.saucedemo.com/ (pagina de login). Los campos de login estan vacios. No se puede acceder a /inventory.html sin volver a loguearse.",
        "Alta"
    ],

    # ==================== SEGURIDAD ====================
    [
        "TC-SEC-001",
        "Seguridad",
        "Acceso directo a URL internas sin autenticacion",
        "1. Cerrar sesion o abrir navegador limpio\n2. Navegar directamente a https://www.saucedemo.com/inventory.html\n3. Verificar que redirecciona a login o muestra error",
        "URL: /inventory.html sin sesion activa",
        "NO debe mostrar el inventario. Debe redireccionar a la pagina de login (https://www.saucedemo.com/). Verificar que no hay acceso a recursos protegidos sin autenticacion.",
        "Alta"
    ],
    [
        "TC-SEC-002",
        "Seguridad",
        "Acceso directo a carrito sin autenticacion",
        "1. Navegar directamente a https://www.saucedemo.com/cart.html\n2. Verificar que NO se accede",
        "URL: /cart.html sin sesion",
        "Redirecciona a login. No muestra informacion del carrito.",
        "Alta"
    ],
    [
        "TC-SEC-003",
        "Seguridad",
        "Acceso directo a checkout sin autenticacion",
        "1. Navegar directamente a https://www.saucedemo.com/checkout-step-one.html\n2. Verificar que NO se accede",
        "URL: /checkout-step-one.html sin sesion",
        "Redirecciona a login. No muestra formulario de checkout.",
        "Alta"
    ],
    [
        "TC-SEC-004",
        "Seguridad",
        "Verificar que password no se muestra en texto plano en el DOM",
        "1. Login como standard_user\n2. Inspeccionar el campo password con herramientas de desarrollador\n3. Verificar que el atributo type='password' esta presente\n4. Verificar que el valor no es visible en el DOM",
        "Usuario: standard_user\nPassword: secret_sauce",
        "El campo password tiene type='password'. Los caracteres se muestran como puntos/bullets. El valor no es legible en el DOM. No hay atributos 'value' expuestos en texto plano.",
        "Alta"
    ],
    [
        "TC-SEC-005",
        "Seguridad",
        "Verificar existencia de HTTPS y certificado valido",
        "1. Navegar a https://www.saucedemo.com/\n2. Verificar que el candado de seguridad aparece en la barra de direcciones\n3. Verificar que el certificado SSL es valido y no expirado",
        "URL: https://www.saucedemo.com/",
        "La conexion debe ser HTTPS. El candado de seguridad debe aparecer. El certificado SSL debe ser valido (no expirado, emisor confiable).",
        "Alta"
    ],
    [
        "TC-SEC-006",
        "Seguridad",
        "Verificar que la sesion se invalida despues de logout (no usar back)",
        "1. Login como standard_user\n2. Hacer logout desde el menu\n3. Hacer clic en 'Back' del navegador\n4. Verificar que NO se accede a la pagina anterior",
        "Usuario: standard_user\nAccion: logout + boton Back del navegador",
        "Al hacer clic en Back, debe redireccionar a login o mostrar pagina expirada. No debe mostrar el inventario ni permitir interaccion sin autenticacion.",
        "Alta"
    ],
    [
        "TC-SEC-007",
        "Seguridad",
        "Verificar que la sesion expira/cierra al cerrar pestana",
        "1. Login como standard_user en pestana A\n2. Copiar URL de inventory.html\n3. Cerrar pestana A\n4. Abrir nueva pestana con la URL copiada\n5. Verificar que redirecciona a login",
        "URL: /inventory.html en nueva pestana despues de cerrar la original",
        "Debe redireccionar a login. La sesion no debe persistir entre pestanas cerradas (dependiendo de la configuracion de sesion del navegador).",
        "Media"
    ],

    # ==================== PERFORMANCE ====================
    [
        "TC-PERF-001",
        "Performance",
        "Tiempo de carga de pagina de login",
        "1. Medir tiempo de carga completo de https://www.saucedemo.com/\n2. Repetir 3 veces\n3. Calcular promedio\n4. Verificar que sea menor a 3 segundos",
        "URL: https://www.saucedemo.com/\nIteraciones: 3",
        "El tiempo de carga promedio debe ser < 3 segundos. La pagina debe ser completamente interactiva (DOM listo, JS cargado).",
        "Media"
    ],
    [
        "TC-PERF-002",
        "Performance",
        "Tiempo de respuesta del login",
        "1. Navegar a login\n2. Iniciar temporizador\n3. Ingresar credenciales y hacer login\n4. Detener temporizador cuando inventory carga\n5. Repetir 3 veces con standard_user y promediar",
        "Usuario: standard_user\nAccion: login y medir tiempo\nIteraciones: 3",
        "Tiempo de respuesta del login debe ser < 2 segundos para standard_user. Para performance_glitch_user se espera mayor latencia (3-6 segundos) pero debe completarse.",
        "Media"
    ],
    [
        "TC-PERF-003",
        "Performance",
        "Tiempo de renderizado del inventario con 6 productos",
        "1. Login como standard_user\n2. Medir tiempo desde que /inventory.html comienza a cargar hasta que todos los 6 productos son visibles e interactivos\n3. Verificar rendimiento",
        "Usuario: standard_user\nMetrica: tiempo de renderizado completo",
        "El inventario debe renderizar completamente en < 2 segundos. Todos los productos deben ser visibles y los botones 'Add to cart' deben ser clicables.",
        "Media"
    ],
    [
        "TC-PERF-004",
        "Performance",
        "Tiempo de respuesta del checkout completo (flujo completo)",
        "1. Login como standard_user\n2. Agregar producto\n3. Ir al carrito\n4. Checkout con datos\n5. Finalizar compra\n6. Medir tiempo total del flujo completo\n7. Repetir 3 veces",
        "Usuario: standard_user\nFlujo: login -> add -> cart -> checkout -> finish\nIteraciones: 3",
        "El flujo completo de compra debe completarse en < 10 segundos total (incluyendo navegacion entre paginas).",
        "Baja"
    ],

    # ==================== LIMITE / BORDE ====================
    [
        "TC-LIM-001",
        "Limite",
        "Agregar cantidad maxima de productos al carrito (todos los 6)",
        "1. Login como standard_user\n2. Agregar los 6 productos disponibles al carrito\n3. Verificar badge muestra '6'\n4. Ir al carrito y verificar que los 6 estan listados",
        "Usuario: standard_user\nProductos: todos los 6 disponibles",
        "Badge del carrito muestra '6'. Los 6 productos aparecen en el carrito con sus respectivos precios. La interfaz no se rompe ni se superpone.",
        "Media"
    ],
    [
        "TC-LIM-002",
        "Limite",
        "Rapid Add and Remove (agregar y quitar rapidamente)",
        "1. Login como standard_user\n2. Rápidamente agregar y remover el mismo producto 10 veces seguidas\n3. Verificar estado final del badge",
        "Usuario: standard_user\nProducto: Sauce Labs Backpack\nAcciones: add/remove rapidos x10",
        "El estado final del carrito debe ser consistente. Si se termina en 'removed', badge no debe aparecer. Si termina en 'added', badge muestra '1'. No debe haber estados inconsistentes.",
        "Media"
    ],
    [
        "TC-LIM-003",
        "Limite",
        "Checkout con codigo postal extremadamente largo",
        "1. Login como standard_user\n2. Agregar producto, ir a checkout\n3. Ingresar codigo postal de 100 caracteres\n4. Completar checkout",
        "Usuario: standard_user\nNombre: Juan\nApellido: Perez\nCP: 'A'*100",
        "El sistema debe aceptar el codigo postal largo o truncarlo. No debe mostrar error de validacion a menos que haya una regla explicita de longitud maxima. El checkout debe completarse.",
        "Baja"
    ],
    [
        "TC-LIM-004",
        "Limite",
        "Checkout con nombres con caracteres especiales y acentos",
        "1. Login como standard_user\n2. Agregar producto, ir a checkout\n3. Ingresar nombres con acentos, caracteres especiales y numeros\n4. Completar checkout",
        "Usuario: standard_user\nNombre: José María O'Brien\nApellido: Müller-Straße 123\nCP: 12345",
        "El sistema debe aceptar caracteres Unicode, acentos, guiones y numeros en los campos de nombre/apellido. El checkout debe completarse exitosamente.",
        "Baja"
    ],

    # ==================== COMPATIBILIDAD NAVEGADORES ====================
    [
        "TC-COMP-001",
        "Compatibilidad",
        "Login y flujo basico en Chrome",
        "1. Abrir Chrome\n2. Navegar a https://www.saucedemo.com/\n3. Login como standard_user\n4. Agregar producto al carrito\n5. Hacer checkout completo\n6. Verificar que todo funciona correctamente en Chrome",
        "Navegador: Chrome (ultima version)\nUsuario: standard_user\nFlujo completo",
        "Todas las funcionalidades funcionan correctamente en Chrome: login, inventario, carrito, checkout, menu, ordenamiento, botones.",
        "Alta"
    ],
    [
        "TC-COMP-002",
        "Compatibilidad",
        "Login y flujo basico en Firefox",
        "1. Abrir Firefox\n2. Navegar a https://www.saucedemo.com/\n3. Login como standard_user\n4. Agregar 2 productos\n5. Hacer checkout\n6. Verificar que todo funciona en Firefox",
        "Navegador: Firefox (ultima version)\nUsuario: standard_user\nFlujo completo",
        "Todas las funcionalidades funcionan correctamente en Firefox. El layout se renderiza correctamente. No hay diferencias visuales significativas vs Chrome.",
        "Alta"
    ],
    [
        "TC-COMP-003",
        "Compatibilidad",
        "Login y flujo basico en Edge",
        "1. Abrir Edge (Chromium)\n2. Navegar a https://www.saucedemo.com/\n3. Login como standard_user\n4. Agregar producto\n5. Checkout completo\n6. Verificar que todo funciona en Edge",
        "Navegador: Edge (ultima version)\nUsuario: standard_user\nFlujo completo",
        "Todas las funcionalidades funcionan correctamente en Edge. Al ser Chromium-based, el comportamiento debe ser similar a Chrome.",
        "Alta"
    ],
    [
        "TC-COMP-004",
        "Compatibilidad",
        "Verificar layout responsivo en resoluciones de escritorio (1920x1080)",
        "1. Abrir navegador con resolucion 1920x1080\n2. Login y navegar inventario\n3. Verificar que los productos se muestran en grilla de 3 columnas\n4. No hay elementos superpuestos ni truncados",
        "Resolucion: 1920x1080 (escritorio Full HD)",
        "Grilla de productos en 3 columnas. Todo el contenido es visible sin scroll horizontal. Los textos no se superponen. Los botones estan alineados.",
        "Media"
    ],
    [
        "TC-COMP-005",
        "Compatibilidad",
        "Verificar layout en resolucion de tablet (768x1024)",
        "1. Abrir navegador con resolucion 768x1024\n2. Login y navegar inventario\n3. Verificar que los productos se reordenan para la resolucion\n4. No hay elementos cortados",
        "Resolucion: 768x1024 (tablet portrait)",
        "Los productos se muestran en 2 columnas (responsive). Todo el contenido es accesible. No hay scroll horizontal. El menu burger funciona correctamente.",
        "Media"
    ],
    [
        "TC-COMP-006",
        "Compatibilidad",
        "Verificar layout en resolucion movil (375x667)",
        "1. Abrir navegador con resolucion 375x667\n2. Login y navegar inventario\n3. Verificar layout responsivo\n4. Verificar que todos los elementos son interactivos en pantalla pequeña",
        "Resolucion: 375x667 (iPhone 6/7/8)",
        "Los productos se muestran en 1 columna. Todo es accesible con scroll vertical. Los botones son lo suficientemente grandes para tocar en movil. No hay elementos rotos.",
        "Media"
    ],
    [
        "TC-COMP-007",
        "Compatibilidad",
        "Verificar que el zoom del navegador no rompe el layout (80% y 150%)",
        "1. Login como standard_user\n2. Aplicar zoom al 80%\n3. Verificar layout\n4. Aplicar zoom al 150%\n5. Verificar layout y funcionalidad",
        "Zoom: 80% y 150%\nNavegador: Chrome",
        "Con zoom al 80%: todo es mas pequeno pero funcional, sin elementos superpuestos. Con zoom al 150%: aparece scroll pero los elementos mantienen su estructura y son interactivos.",
        "Baja"
    ],

    # ==================== CONTINUIDAD / ESTADO ====================
    [
        "TC-CONT-001",
        "Continuidad",
        "Persistencia del carrito al navegar entre paginas",
        "1. Login como standard_user\n2. Agregar producto en inventario\n3. Ir al detalle de otro producto\n4. Volver al inventario\n5. Ir al carrito\n6. Verificar que el producto agregado persiste",
        "Usuario: standard_user\nNavegacion: inventory -> detail -> inventory -> cart",
        "El producto agregado inicialmente permanece en el carrito a traves de toda la navegacion. El badge se mantiene consistente.",
        "Alta"
    ],
    [
        "TC-CONT-002",
        "Continuidad",
        "Persistencia del carrito al hacer logout y login",
        "1. Login como standard_user\n2. Agregar productos al carrito\n3. Hacer logout\n4. Login nuevamente\n5. Ir al carrito\n6. Verificar si los productos persisten",
        "Usuario: standard_user\nAccion: add -> logout -> login -> check cart",
        "Despues de logout y login, el carrito debe estar vacio (la sesion anterior se pierde). La sesion no debe persistir entre logins.",
        "Media"
    ],
    [
        "TC-CONT-003",
        "Continuidad",
        "Uso del boton Back del navegador durante el checkout",
        "1. Login como standard_user\n2. Agregar producto\n3. Ir al carrito\n4. Clic en Checkout (step-one)\n5. Presionar Back del navegador\n6. Verificar que vuelve al carrito\n7. Presionar Back otra vez y verificar que vuelve al inventario",
        "Usuario: standard_user\nNavegacion: cart -> checkout-step-one, luego Back 2 veces",
        "Primer Back: vuelve al carrito con los productos intactos. Segundo Back: vuelve al inventario. No hay pantallas en blanco ni errores de javascript.",
        "Media"
    ],
]

for row_data in test_cases:
    ws.append(row_data)

priority_fills = {
    "Alta": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "Media": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "Baja": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

type_fills = {
    "Positivo": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "Negativo": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "Seguridad": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
    "Performance": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "Limite": PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"),
    "Compatibilidad": PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid"),
    "Continuidad": PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid"),
}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    priority_val = str(row[6].value) if row[6].value else ""
    type_val = str(row[1].value) if row[1].value else ""
    if priority_val in priority_fills:
        row[6].fill = priority_fills[priority_val]
    if type_val in type_fills:
        row[1].fill = type_fills[type_val]
    for cell in row:
        cell.alignment = wrap_align
        cell.border = thin_border

col_widths = [14, 16, 55, 70, 40, 65, 10]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.auto_filter.ref = f"A1:G{ws.max_row}"
ws.freeze_panes = "A2"

ws.row_dimensions[1].height = 30
for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 120

filepath = "Casos_de_Prueba_Saucedemo.xlsx"
wb.save(filepath)
print(f"Documento generado exitosamente: {filepath}")
print(f"Total de casos de prueba: {len(test_cases)}")
